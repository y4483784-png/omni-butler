"""Tabular schema inference for xlsx/csv analysis planning."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

_MAX_HINT_COLS = 30
_MAX_PROFILE_VALUES = 5
_MAX_SHEETS = 12
_DATE_HINT_RE = re.compile(r"(date|time|日期|时间|年月|月份|日)", re.I)
_ID_HINT_RE = re.compile(r"(id|编号|序号|学号|工号|code|编码)", re.I)
_TEXT_DIM_HINT_RE = re.compile(r"(name|label|title|type|category)", re.I)


@dataclass
class SheetProfile:
    name: str
    row_count: int
    columns: list[str] = field(default_factory=list)
    numeric_candidates: list[str] = field(default_factory=list)
    dimension_candidates: list[str] = field(default_factory=list)
    date_candidates: list[str] = field(default_factory=list)
    sample_values: dict[str, list[str]] = field(default_factory=dict)
    null_rates: dict[str, float] = field(default_factory=dict)
    cardinalities: dict[str, int] = field(default_factory=dict)
    rough_quantiles: dict[str, dict[str, float]] = field(default_factory=dict)


@dataclass
class TabularSchema:
    filename: str
    file_type: str
    sheets: list[SheetProfile] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)
    dimension_candidates: list[str] = field(default_factory=list)
    measure_candidates: list[str] = field(default_factory=list)
    date_candidates: list[str] = field(default_factory=list)
    aliases: dict[str, str] = field(default_factory=dict)
    unreadable_reason: str = ""

    @property
    def is_readable(self) -> bool:
        return not self.unreadable_reason


def infer_tabular_schema(path: str | Path) -> TabularSchema:
    path = Path(path)
    if not path.is_file():
        return TabularSchema(filename=path.name, file_type=path.suffix.lower(), unreadable_reason="文件不可读")
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _infer_xlsx(path)
    if ext == ".csv":
        return _infer_csv(path)
    return TabularSchema(filename=path.name, file_type=ext, unreadable_reason=f"不支持的表格类型：{ext}")


def tabular_file_hints(path: str | Path) -> str:
    return schema_to_hints(infer_tabular_schema(path))


def schema_to_hints(schema: TabularSchema) -> str:
    head = f"文件：{schema.filename}（{schema.file_type.upper().lstrip('.')}）"
    if not schema.is_readable:
        return f"{head}\n  {schema.unreadable_reason}"

    lines = [head]
    if schema.columns:
        lines.append(f"  全局列：{', '.join(schema.columns[:_MAX_HINT_COLS])}")
    if schema.dimension_candidates:
        lines.append(f"  候选维度列：{', '.join(schema.dimension_candidates[:12])}")
    if schema.measure_candidates:
        lines.append(f"  候选数值列：{', '.join(schema.measure_candidates[:12])}")
    if schema.date_candidates:
        lines.append(f"  候选日期列：{', '.join(schema.date_candidates[:12])}")
    if schema.aliases:
        alias_text = ", ".join(f"{k}->{v}" for k, v in list(schema.aliases.items())[:8])
        lines.append(f"  常见别名：{alias_text}")

    for sheet in schema.sheets[:_MAX_SHEETS]:
        lines.append(
            f"  sheet「{sheet.name}」约 {sheet.row_count} 数据行，表头："
            f"{', '.join(sheet.columns[:_MAX_HINT_COLS]) if sheet.columns else '(无)'}"
        )
        if sheet.null_rates:
            top_null = sorted(sheet.null_rates.items(), key=lambda x: -x[1])[:5]
            bits = [f"{c}:{r:.0%}" for c, r in top_null if r > 0]
            if bits:
                lines.append(f"    空值率：{', '.join(bits)}")
        if sheet.rough_quantiles:
            for col, qs in list(sheet.rough_quantiles.items())[:3]:
                lines.append(
                    f"    {col} 粗分位 p25={qs.get('p25')} p50={qs.get('p50')} p75={qs.get('p75')}"
                )
    return "\n".join(lines)


def _normalize_col(name: object) -> str:
    text = str(name or "").strip()
    return re.sub(r"\s+", "", text)


def _merge_unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


def _csv_rows(path: Path) -> tuple[list[str], list[list[str]], int]:
    for enc in ("utf-8-sig", "utf-8", "gbk", "latin-1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                raw_header = next(reader, None) or []
                rows = [row for row in reader]
            cols = [_normalize_col(c) for c in raw_header if _normalize_col(c)]
            return cols, rows, len(rows)
        except UnicodeDecodeError:
            continue
    return [], [], 0


def _infer_csv(path: Path) -> TabularSchema:
    cols, rows, row_count = _csv_rows(path)
    if not cols:
        return TabularSchema(filename=path.name, file_type=".csv", unreadable_reason="CSV 未解析到列名")
    profile = _profile_rows("CSV", cols, rows)
    return TabularSchema(
        filename=path.name,
        file_type=".csv",
        sheets=[profile],
        columns=profile.columns,
        dimension_candidates=profile.dimension_candidates,
        measure_candidates=profile.numeric_candidates,
        date_candidates=profile.date_candidates,
        aliases={},
    )


def _infer_xlsx(path: Path) -> TabularSchema:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return TabularSchema(filename=path.name, file_type=".xlsx", unreadable_reason="缺少 openpyxl")

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        return TabularSchema(filename=path.name, file_type=".xlsx", unreadable_reason=f"打开失败：{e}")

    profiles: list[SheetProfile] = []
    all_cols: list[str] = []
    dims: list[str] = []
    nums: list[str] = []
    dates: list[str] = []
    try:
        for sn in wb.sheetnames[:_MAX_SHEETS]:
            ws = wb[sn]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                profiles.append(SheetProfile(name=sn, row_count=0))
                continue
            cols = [_normalize_col(c) for c in (header_row or ()) if _normalize_col(c)]
            rows: list[list[str]] = []
            row_count = 0
            for raw in rows_iter:
                vals = ["" if c is None else str(c).strip() for c in (raw or ())]
                if any(v for v in vals):
                    row_count += 1
                rows.append(vals)
            profile = _profile_rows(sn, cols, rows, row_count=row_count)
            profiles.append(profile)
            all_cols.extend(profile.columns)
            dims.extend(profile.dimension_candidates)
            nums.extend(profile.numeric_candidates)
            dates.extend(profile.date_candidates)
    finally:
        wb.close()

    merged_cols = _merge_unique(all_cols)
    return TabularSchema(
        filename=path.name,
        file_type=".xlsx",
        sheets=profiles,
        columns=merged_cols,
        dimension_candidates=_merge_unique(dims),
        measure_candidates=_merge_unique(nums),
        date_candidates=_merge_unique(dates),
        aliases={},
    )


def _profile_rows(sheet_name: str, cols: list[str], rows: list[list[str]], row_count: int | None = None) -> SheetProfile:
    non_empty_rows = 0
    numeric_hits = {c: 0 for c in cols}
    date_hits = {c: 0 for c in cols}
    null_counts = {c: 0 for c in cols}
    text_unique: dict[str, set[str]] = {c: set() for c in cols}
    sample_values: dict[str, list[str]] = {c: [] for c in cols}
    numeric_vals: dict[str, list[float]] = {c: [] for c in cols}

    for row in rows:
        if any(str(v or "").strip() for v in row):
            non_empty_rows += 1
        for idx, col in enumerate(cols):
            raw = row[idx] if idx < len(row) else ""
            text = str(raw or "").strip()
            if not text:
                null_counts[col] += 1
                continue
            if len(sample_values[col]) < _MAX_PROFILE_VALUES and text not in sample_values[col]:
                sample_values[col].append(text)
            if _looks_numeric(text):
                numeric_hits[col] += 1
                try:
                    numeric_vals[col].append(float(text.replace(",", "").replace("%", "")))
                except ValueError:
                    pass
            elif _looks_date(text):
                date_hits[col] += 1
            else:
                if len(text_unique[col]) < 50:
                    text_unique[col].add(text)

    real_rows = row_count if row_count is not None else max(non_empty_rows, 1)
    numeric_candidates = [
        c for c in cols if numeric_hits[c] >= max(1, min(real_rows, 3))
    ]
    date_candidates = [
        c
        for c in cols
        if _DATE_HINT_RE.search(c) or date_hits[c] >= max(1, min(real_rows, 3))
    ]
    dimension_candidates: list[str] = []
    for c in cols:
        unique_count = len(text_unique[c])
        if c in numeric_candidates:
            continue
        if _ID_HINT_RE.search(c):
            continue
        if c in date_candidates:
            continue
        if _TEXT_DIM_HINT_RE.search(c) or 2 <= unique_count <= max(3, min(real_rows, 30)):
            dimension_candidates.append(c)

    null_rates = {c: round(null_counts[c] / max(real_rows, 1), 3) for c in cols}
    cardinalities = {
        c: len(text_unique[c]) + (1 if numeric_hits[c] else 0) for c in cols
    }
    rough_quantiles: dict[str, dict[str, float]] = {}
    for c, vals in numeric_vals.items():
        if len(vals) < 3:
            continue
        sorted_v = sorted(vals)
        n = len(sorted_v)

        def _q(p: float) -> float:
            i = min(n - 1, max(0, int(round((n - 1) * p))))
            return round(sorted_v[i], 4)

        rough_quantiles[c] = {"p25": _q(0.25), "p50": _q(0.5), "p75": _q(0.75)}

    return SheetProfile(
        name=sheet_name,
        row_count=real_rows if row_count is not None else non_empty_rows,
        columns=cols,
        numeric_candidates=_merge_unique(numeric_candidates),
        dimension_candidates=_merge_unique(dimension_candidates),
        date_candidates=_merge_unique(date_candidates),
        sample_values={k: v for k, v in sample_values.items() if v},
        null_rates=null_rates,
        cardinalities=cardinalities,
        rough_quantiles=rough_quantiles,
    )


def _looks_date(text: str) -> bool:
    t = text.strip()
    if not t or len(t) > 32:
        return False
    if re.match(r"^\d{4}[-/年]\d{1,2}([-/月]\d{1,2})?", t):
        return True
    if re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$", t):
        return True
    return False


def _looks_numeric(text: str) -> bool:
    t = text.strip().replace(",", "")
    if t.endswith("%"):
        t = t[:-1]
    if not t or len(t) > 40:
        return False
    try:
        float(t)
        return True
    except ValueError:
        return False
